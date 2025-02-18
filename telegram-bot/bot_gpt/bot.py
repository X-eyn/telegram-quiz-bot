import logging
import json
import os
import csv
import asyncio
from datetime import datetime
from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    Poll,
    KeyboardButton,
    KeyboardButtonPollType,
    ReplyKeyboardMarkup,
    BotCommand,
)
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    ConversationHandler,
    PollAnswerHandler,
    filters,
    ContextTypes,
)

# -------------------------------
# Conversation States
# -------------------------------
# Quiz Creation states
CR_QUIZ_NAME = 1
CR_QUIZ_DESCRIPTION = 2
CR_PRE_QUESTION = 3
CR_ADD_QUESTION = 4
CR_NEXT_QUESTION = 5
CR_TIMER = 6
CR_SHUFFLE = 7
CR_SUBMIT = 8
CR_QUESTION_IMAGE = 9  # New state for optional question image

# Quiz Taking states
TK_SELECT_QUIZ = 19
TK_READY = 20       # New state: after selecting a quiz, show ready message
TK_RUNNING = 21     # Quiz is running (auto-advance after each answer)

# Import Conversation state
IMPORT_WAITING_FILE = 30

# -------------------------------
# File names for persistent storage
# -------------------------------
QUIZZES_FILE = "quizzes.json"
PARTICIPANTS_FILE = "quiz_participants.json"

# -------------------------------
# Default Quiz (used if no quiz has been created)
# -------------------------------
DEFAULT_QUIZ = {
    "quiz_name": "Default Sample Quiz",
    "quiz_description": "This is a default quiz. Create your own quiz using the bot!",
    "pre_question": "",
    "questions": [
        {
            "question": "What is the capital of France?",
            "options": ["Paris", "Rome", "Madrid", "Berlin"],
            "correct_option_id": 0,
            "explanation": "Paris is the capital of France.",
        },
        {
            "question": "What is 2 + 2?",
            "options": ["3", "4", "5"],
            "correct_option_id": 1,
            "explanation": "2 + 2 equals 4.",
        },
    ],
    "timer": 15,
    "shuffle": False,
    "submit": False,
}

# -------------------------------
# Helper functions for JSON storage
# -------------------------------
def load_quizzes():
    if os.path.exists(QUIZZES_FILE):
        with open(QUIZZES_FILE, "r", encoding="utf-8") as f:
            try:
                data = json.load(f)
                if isinstance(data, list):
                    return data
                else:
                    return []
            except json.JSONDecodeError:
                return []
    else:
        return []


def save_quiz_data(quiz):
    quizzes = load_quizzes()
    quizzes.append(quiz)
    with open(QUIZZES_FILE, "w", encoding="utf-8") as f:
        json.dump(quizzes, f, indent=2)


def save_poll_answer_data(data):
    if os.path.exists(PARTICIPANTS_FILE):
        with open(PARTICIPANTS_FILE, "r", encoding="utf-8") as f:
            try:
                all_data = json.load(f)
            except json.JSONDecodeError:
                all_data = []
    else:
        all_data = []
    all_data.append(data)
    with open(PARTICIPANTS_FILE, "w", encoding="utf-8") as f:
        json.dump(all_data, f, indent=2)


# -------------------------------
# /start Command – Main Menu
# -------------------------------
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    keyboard = [
        [InlineKeyboardButton("Create Quiz", callback_data="create_quiz")],
        [InlineKeyboardButton("Take Quiz", callback_data="take_quiz")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    if update.message:
        await update.message.reply_text(
            "Welcome to the Interactive Quiz Bot!\nPlease choose an option:",
            reply_markup=reply_markup,
        )
    elif update.callback_query:
        await update.callback_query.edit_message_text(
            "Welcome to the Interactive Quiz Bot!\nPlease choose an option:",
            reply_markup=reply_markup,
        )


# -------------------------------
# QUIZ CREATION FLOW – Using Native Polls for Questions (with optional images)
# -------------------------------
async def cr_entry(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    context.user_data.clear()
    context.user_data['questions'] = []
    await query.edit_message_text("Let's create your quiz!\n\nPlease enter the quiz name:")
    return CR_QUIZ_NAME


async def cr_quiz_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data['quiz_name'] = update.message.text
    await update.message.reply_text("Enter a description for your quiz (or type /skip to skip):")
    return CR_QUIZ_DESCRIPTION


async def cr_quiz_description(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data['quiz_description'] = update.message.text
    await update.message.reply_text("Send pre-question text or image (optional, or type /skip to skip):")
    return CR_PRE_QUESTION


async def cr_skip_description(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data['quiz_description'] = ""
    await update.message.reply_text("No description added.\nSend pre-question text or image (optional, or type /skip to skip):")
    return CR_PRE_QUESTION


async def cr_pre_question(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if update.effective_chat.type != "private":
        await update.message.reply_text(
            "Poll creation via button is only available in private chats. "
            "Please send your quiz poll manually (as a quiz-type poll) to the chat."
        )
        return CR_ADD_QUESTION
    if update.message.photo:
        context.user_data['pre_question'] = {"type": "photo", "file_id": update.message.photo[-1].file_id}
    else:
        context.user_data['pre_question'] = {"type": "text", "content": update.message.text}
    button = KeyboardButton("Create Quiz Question", request_poll=KeyboardButtonPollType(type="quiz"))
    markup = ReplyKeyboardMarkup([[button]], one_time_keyboard=True, resize_keyboard=True)
    await update.message.reply_text(
        "Now, create your first quiz question by pressing the button below:",
        reply_markup=markup,
    )
    return CR_ADD_QUESTION


async def cr_skip_pre_question(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if update.effective_chat.type != "private":
        await update.message.reply_text(
            "Poll creation via button is only available in private chats. Please send your quiz poll manually (as a quiz-type poll) to the chat."
        )
        return CR_ADD_QUESTION
    context.user_data['pre_question'] = None
    button = KeyboardButton("Create Quiz Question", request_poll=KeyboardButtonPollType(type="quiz"))
    markup = ReplyKeyboardMarkup([[button]], one_time_keyboard=True, resize_keyboard=True)
    await update.message.reply_text(
        "No pre-question media/text added.\nNow, create your first quiz question by pressing the button below:",
        reply_markup=markup,
    )
    return CR_ADD_QUESTION


# Modified to store poll data temporarily and prompt for an optional image.
async def handle_created_poll(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    poll = update.message.poll
    if poll is None:
        return CR_ADD_QUESTION
    if poll.type != "quiz":
        await update.message.reply_text("Please create a quiz-type poll.")
        return CR_ADD_QUESTION
    question_data = {
        "question": poll.question,
        "options": [option.text for option in poll.options],
        "correct_option_id": poll.correct_option_id,
        "explanation": poll.explanation or ""
    }
    # Store temporarily for optional image attachment
    context.user_data['current_question'] = question_data
    await update.message.reply_text("Would you like to attach an image to this question? If yes, please send the image now. Otherwise, type /skip.")
    return CR_QUESTION_IMAGE


# Handler when a photo is sent in CR_QUESTION_IMAGE state.
async def cr_add_question_image(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    photo_file_id = update.message.photo[-1].file_id
    question_data = context.user_data.get('current_question')
    if question_data is None:
        await update.message.reply_text("No question found. Please send a quiz poll first.")
        return CR_ADD_QUESTION
    question_data['image'] = photo_file_id
    context.user_data.setdefault('questions', []).append(question_data)
    context.user_data.pop('current_question', None)
    keyboard = [
        [InlineKeyboardButton("Add another question", callback_data="add_another")],
        [InlineKeyboardButton("Finish questions", callback_data="finish_questions")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("Question added with image. What would you like to do next?", reply_markup=reply_markup)
    return CR_NEXT_QUESTION


# Handler for skipping image addition.
async def cr_skip_question_image(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    question_data = context.user_data.pop('current_question', None)
    if question_data:
        context.user_data.setdefault('questions', []).append(question_data)
    keyboard = [
        [InlineKeyboardButton("Add another question", callback_data="add_another")],
        [InlineKeyboardButton("Finish questions", callback_data="finish_questions")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("Question added without image. What would you like to do next?", reply_markup=reply_markup)
    return CR_NEXT_QUESTION


async def cr_next_question(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    if query.data == "add_another":
        if update.effective_chat.type != "private":
            await query.edit_message_text("Poll creation via button is only available in private chats. Please send your quiz poll manually (as a quiz-type poll) to the chat.")
            return CR_ADD_QUESTION
        button = KeyboardButton("Create Quiz Question", request_poll=KeyboardButtonPollType(type="quiz"))
        markup = ReplyKeyboardMarkup([[button]], one_time_keyboard=True, resize_keyboard=True)
        await query.edit_message_text("Send your next quiz question by pressing the button below:")
        await query.message.reply_text("Press the button below to create a new quiz question:", reply_markup=markup)
        return CR_ADD_QUESTION
    elif query.data == "finish_questions":
        await query.edit_message_text("Enter the timer duration (in seconds) for each question:")
        return CR_TIMER


async def cr_finish_questions(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not context.user_data.get('questions'):
        await update.message.reply_text("You haven't added any questions yet. Please add at least one question using the poll button.")
        return CR_ADD_QUESTION
    await update.message.reply_text("Enter the timer duration (in seconds) for each question:")
    return CR_TIMER


async def cr_set_timer(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        context.user_data['timer'] = int(update.message.text)
    except ValueError:
        await update.message.reply_text("Please enter a valid number for the timer (in seconds):")
        return CR_TIMER
    keyboard = [
        [InlineKeyboardButton("Yes", callback_data="shuffle_yes"),
         InlineKeyboardButton("No", callback_data="shuffle_no")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("Do you want to shuffle questions and answer options?", reply_markup=reply_markup)
    return CR_SHUFFLE


async def cr_set_shuffle(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    context.user_data['shuffle'] = (query.data == "shuffle_yes")
    keyboard = [
        [InlineKeyboardButton("Yes", callback_data="submit_yes"),
         InlineKeyboardButton("No", callback_data="submit_no")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text("Would you like to submit your quiz to the contest?", reply_markup=reply_markup)
    return CR_SUBMIT


async def cr_set_submit(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    context.user_data['submit'] = (query.data == "submit_yes")
    quiz = {
        "quiz_name": context.user_data.get('quiz_name', 'Untitled Quiz'),
        "quiz_description": context.user_data.get('quiz_description', ''),
        "pre_question": context.user_data.get('pre_question', None),
        "questions": context.user_data.get('questions', []),
        "timer": context.user_data.get('timer', 15),
        "shuffle": context.user_data.get('shuffle', False),
        "submit": context.user_data.get('submit', False)
    }
    summary = (
        f"Quiz Created!\n\n"
        f"Title: {quiz['quiz_name']}\n"
        f"Description: {quiz['quiz_description']}\n"
        f"Pre-question: {quiz['pre_question']}\n"
        f"Number of questions: {len(quiz['questions'])}\n"
        f"Timer: {quiz['timer']} seconds\n"
        f"Shuffle: {'Yes' if quiz['shuffle'] else 'No'}\n"
        f"Submitted to contest: {'Yes' if quiz['submit'] else 'No'}"
    )
    await query.edit_message_text(summary)
    save_quiz_data(quiz)
    await query.message.reply_text("Your quiz has been created! You can take it by selecting 'Take Quiz' from the main menu.")
    return ConversationHandler.END


async def cr_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text("Quiz creation cancelled.")
    return ConversationHandler.END


# -------------------------------
# IMPORT QUIZZES – Import from JSON/CSV/XLSX
# -------------------------------
async def import_quizzes_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text("Please send a document file (JSON, CSV, or XLSX) containing quiz data to import.")
    return IMPORT_WAITING_FILE


async def import_quizzes_file(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    document = update.message.document
    if not document:
        await update.message.reply_text("No document received. Please send a valid document file.")
        return IMPORT_WAITING_FILE
    file_name = document.file_name.lower() if document.file_name else ""
    file = await document.get_file()
    temp_path = f"temp_{file_name}"
    await file.download_to_drive(temp_path)
    imported_quizzes = []
    try:
        if file_name.endswith(".json"):
            with open(temp_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, list):
                    imported_quizzes = data
                else:
                    imported_quizzes = [data]
        elif file_name.endswith(".csv"):
            with open(temp_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                quizzes_dict = {}
                for row in reader:
                    quiz_name = row.get("quiz_name", "Untitled Quiz")
                    if quiz_name not in quizzes_dict:
                        quizzes_dict[quiz_name] = {
                            "quiz_name": quiz_name,
                            "quiz_description": row.get("quiz_description", ""),
                            "pre_question": row.get("pre_question", ""),
                            "questions": [],
                            "timer": int(row.get("timer", 15)),
                            "shuffle": row.get("shuffle", "False") == "True",
                            "submit": row.get("submit", "False") == "True"
                        }
                    question = {
                        "question": row.get("question", ""),
                        "options": row.get("options", "").split(";"),
                        "correct_option_id": int(row.get("correct_option_id", 0)),
                        "explanation": row.get("explanation", "")
                    }
                    if "image" in row and row.get("image", "").strip():
                        question["image"] = row.get("image", "").strip()
                    quizzes_dict[quiz_name]["questions"].append(question)
                imported_quizzes = list(quizzes_dict.values())
        elif file_name.endswith(".xlsx"):
            from openpyxl import load_workbook
            wb = load_workbook(filename=temp_path)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            quizzes_dict = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                quiz_name = row_dict.get("quiz_name", "Untitled Quiz")
                if quiz_name not in quizzes_dict:
                    quizzes_dict[quiz_name] = {
                        "quiz_name": quiz_name,
                        "quiz_description": row_dict.get("quiz_description", ""),
                        "pre_question": row_dict.get("pre_question", ""),
                        "questions": [],
                        "timer": int(row_dict.get("timer", 15)) if row_dict.get("timer") else 15,
                        "shuffle": str(row_dict.get("shuffle", "False")) == "True",
                        "submit": str(row_dict.get("submit", "False")) == "True"
                    }
                question = {
                    "question": row_dict.get("question", ""),
                    "options": row_dict.get("options", "").split(";") if row_dict.get("options") else [],
                    "correct_option_id": int(row_dict.get("correct_option_id", 0)),
                    "explanation": row_dict.get("explanation", "")
                }
                if row_dict.get("image"):
                    question["image"] = row_dict.get("image")
                quizzes_dict[quiz_name]["questions"].append(question)
            imported_quizzes = list(quizzes_dict.values())
        else:
            await update.message.reply_text("Unsupported file type. Please send a JSON, CSV, or XLSX file.")
            os.remove(temp_path)
            return ConversationHandler.END
    except Exception as e:
        await update.message.reply_text(f"Error processing file: {e}")
        os.remove(temp_path)
        return ConversationHandler.END
    os.remove(temp_path)
    existing_quizzes = load_quizzes()
    existing_quizzes.extend(imported_quizzes)
    with open(QUIZZES_FILE, "w", encoding="utf-8") as f:
        json.dump(existing_quizzes, f, indent=2)
    await update.message.reply_text(f"Successfully imported {len(imported_quizzes)} quizzes!")
    return ConversationHandler.END


async def import_quizzes_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text("Import cancelled.")
    return ConversationHandler.END


# -------------------------------
# EXPORT QUIZZES – Export quizzes to JSON, CSV, and XLSX
# -------------------------------
async def export_quizzes_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    quizzes = load_quizzes()
    if not quizzes:
        await update.message.reply_text("No quizzes found to export.")
        return

    json_file = QUIZZES_FILE
    csv_file = "quizzes_export.csv"
    with open(csv_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["quiz_name", "quiz_description", "pre_question", "question", "options", "correct_option_id", "explanation", "image", "timer", "shuffle", "submit"])
        for quiz in quizzes:
            pre_question = ""
            if isinstance(quiz.get("pre_question"), dict):
                if quiz["pre_question"].get("type") == "text":
                    pre_question = quiz["pre_question"].get("content", "")
                elif quiz["pre_question"].get("type") == "photo":
                    pre_question = "[PHOTO]"
            else:
                pre_question = quiz.get("pre_question", "")
            for question in quiz.get("questions", []):
                writer.writerow([
                    quiz.get("quiz_name", ""),
                    quiz.get("quiz_description", ""),
                    pre_question,
                    question.get("question", ""),
                    ";".join(question.get("options", [])),
                    question.get("correct_option_id", 0),
                    question.get("explanation", ""),
                    question.get("image", ""),
                    quiz.get("timer", 15),
                    quiz.get("shuffle", False),
                    quiz.get("submit", False)
                ])
    xlsx_file = "quizzes_export.xlsx"
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["quiz_name", "quiz_description", "pre_question", "question", "options", "correct_option_id", "explanation", "image", "timer", "shuffle", "submit"])
        for quiz in quizzes:
            pre_question = ""
            if isinstance(quiz.get("pre_question"), dict):
                if quiz["pre_question"].get("type") == "text":
                    pre_question = quiz["pre_question"].get("content", "")
                elif quiz["pre_question"].get("type") == "photo":
                    pre_question = "[PHOTO]"
            else:
                pre_question = quiz.get("pre_question", "")
            for question in quiz.get("questions", []):
                ws.append([
                    quiz.get("quiz_name", ""),
                    quiz.get("quiz_description", ""),
                    pre_question,
                    question.get("question", ""),
                    ";".join(question.get("options", [])),
                    question.get("correct_option_id", 0),
                    question.get("explanation", ""),
                    question.get("image", ""),
                    quiz.get("timer", 15),
                    quiz.get("shuffle", False),
                    quiz.get("submit", False)
                ])
        wb.save(xlsx_file)
    except ImportError:
        xlsx_file = None

    chat_id = update.effective_chat.id
    if os.path.exists(json_file):
        await context.bot.send_document(chat_id=chat_id, document=open(json_file, "rb"), filename=json_file)
    if os.path.exists(csv_file):
        await context.bot.send_document(chat_id=chat_id, document=open(csv_file, "rb"), filename=csv_file)
    if xlsx_file and os.path.exists(xlsx_file):
        await context.bot.send_document(chat_id=chat_id, document=open(xlsx_file, "rb"), filename=xlsx_file)

    if os.path.exists(csv_file):
        os.remove(csv_file)
    if xlsx_file and os.path.exists(xlsx_file):
        os.remove(xlsx_file)


# -------------------------------
# QUIZ TAKING FLOW – Using Native Quiz Polls with Timer, Ready Message, and Auto-advance
# -------------------------------
# Entry: load quiz list for selection.
async def tk_entry(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    quizzes = load_quizzes()
    if not quizzes:
        quiz = DEFAULT_QUIZ
        context.user_data['quiz'] = quiz
        context.user_data['question_index'] = 0
        context.user_data['score'] = 0
        # If no quizzes exist, immediately start the quiz
        await query.edit_message_text(f"Starting quiz: {quiz['quiz_name']}\n{quiz['quiz_description']}")
        return await tk_send_poll(query, context)
    else:
        keyboard = []
        for idx, quiz in enumerate(quizzes):
            if isinstance(quiz, dict):
                quiz_name = quiz.get("quiz_name", "Untitled Quiz")
            else:
                quiz_name = str(quiz)
            keyboard.append([InlineKeyboardButton(quiz_name, callback_data=f"select_quiz_{idx}")])
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text("Select a quiz to take:", reply_markup=reply_markup)
        return TK_SELECT_QUIZ


# When a quiz is selected, show a ready screen with quiz details.
async def tk_select_quiz(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    data = query.data
    idx = int(data.split("_")[-1])
    quizzes = load_quizzes()
    if idx < len(quizzes):
        quiz = quizzes[idx]
    else:
        quiz = DEFAULT_QUIZ
    context.user_data['quiz'] = quiz
    context.user_data['question_index'] = 0
    context.user_data['score'] = 0
    # Build the ready message with quiz details.
    num_questions = len(quiz.get("questions", []))
    timer = quiz.get("timer", 15)
    ready_text = (
        f"Get ready for '{quiz['quiz_name']}'!\n"
        f"Number of questions: {num_questions}\n"
        f"Timer per question: {timer} seconds\n\n"
        "Tap 'Ready' to begin or send /cancel to exit."
    )
    ready_button = InlineKeyboardMarkup([[InlineKeyboardButton("Ready", callback_data="ready")]])
    await query.edit_message_text(ready_text, reply_markup=ready_button)
    return TK_READY


# Handler for the "Ready" button. A 3-second countdown is shown before starting the quiz.
async def tk_ready_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    chat_id = query.message.chat_id
    # Send a countdown message
    await context.bot.send_message(chat_id=chat_id, text="Starting in 3...")
    for i in range(2, 0, -1):
        await asyncio.sleep(1)
        await context.bot.send_message(chat_id=chat_id, text=f"Starting in {i}...")
    await asyncio.sleep(1)
    # Start the quiz by sending the first question.
    return await tk_send_poll(query, context)


# Modified tk_send_poll with a fix to handle CallbackQuery objects.
async def tk_send_poll(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    quiz = context.user_data['quiz']
    index = context.user_data['question_index']
    # Determine chat_id either from update or stored data.
    chat_id = None
    if update is not None:
        if hasattr(update, 'effective_chat') and update.effective_chat:
            chat_id = update.effective_chat.id
        elif hasattr(update, 'message') and update.message:
            chat_id = update.message.chat.id
        else:
            # If update is a CallbackQuery, use update.message.chat.id
            if hasattr(update, 'message') and update.message:
                chat_id = update.message.chat.id
            else:
                chat_id = context.user_data.get('quiz_chat_id')
    else:
        chat_id = context.user_data.get('quiz_chat_id')
    
    if not chat_id:
         chat_id = context.user_data.get('quiz_chat_id')
    
    if index < len(quiz['questions']):
        question = quiz['questions'][index]
        # If this question has an image, send it first.
        if "image" in question and question["image"]:
            await context.bot.send_photo(chat_id=chat_id, photo=question["image"])
        options = question.get('options', [])
        correct_option_id = question.get('correct_option_id', 0)
        if correct_option_id is None or not isinstance(correct_option_id, int) or correct_option_id < 0 or correct_option_id >= len(options):
            correct_option_id = 0
        poll_message = await context.bot.send_poll(
            chat_id=chat_id,
            question=question['question'],
            options=options,
            type=Poll.QUIZ,
            correct_option_id=correct_option_id,
            explanation=question.get('explanation', ''),
            is_anonymous=False,
            open_period=quiz.get('timer', 15)
        )
        context.user_data['current_poll_id'] = poll_message.poll.id
        context.user_data['quiz_chat_id'] = chat_id
        return TK_RUNNING
    else:
        score = context.user_data['score']
        total = len(quiz['questions'])
        await context.bot.send_message(
            chat_id=chat_id,
            text=f"Quiz finished!\nYour score: {score}/{total}\nType /start to return to the main menu."
        )
        return ConversationHandler.END


# Global PollAnswer Handler – Save Participant Data, Provide Feedback, and Auto-advance
async def handle_poll_answer(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    poll_answer = update.poll_answer
    if 'quiz' not in context.user_data:
        return
    if context.user_data.get('current_poll_id') != poll_answer.poll_id:
        return
    if not poll_answer.option_ids:
        return
    selected = poll_answer.option_ids[0]
    quiz = context.user_data['quiz']
    index = context.user_data['question_index']
    question = quiz['questions'][index]
    options = question.get('options', [])
    correct_option_id = question.get('correct_option_id', 0)
    if correct_option_id is None or not isinstance(correct_option_id, int) or correct_option_id < 0 or correct_option_id >= len(options):
        correct_option_id = 0
    if selected == correct_option_id:
        context.user_data['score'] += 1
        feedback = "Correct!"
    else:
        correct_option = options[correct_option_id]
        feedback = f"Incorrect. The correct answer was: {correct_option}"
    if question.get('explanation'):
        feedback += f"\nExplanation: {question['explanation']}"
    answer_data = {
        "user_id": poll_answer.user.id,
        "username": poll_answer.user.username,
        "poll_id": poll_answer.poll_id,
        "selected_option": selected,
        "question": question['question'],
        "correct_option": options[correct_option_id],
        "timestamp": datetime.utcnow().isoformat()
    }
    save_poll_answer_data(answer_data)
    quiz_chat_id = context.user_data.get('quiz_chat_id')
    if not quiz_chat_id:
        quiz_chat_id = poll_answer.user.id
    await context.bot.send_message(chat_id=quiz_chat_id, text=feedback)
    context.user_data['current_poll_id'] = None
    # Wait 3 seconds before automatically sending the next question.
    await asyncio.sleep(3)
    context.user_data['question_index'] += 1
    await tk_send_poll(None, context)


async def tk_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text("Quiz cancelled. Use /start to return to the main menu.")
    return ConversationHandler.END


# -------------------------------
# Startup function to set bot commands
# -------------------------------
async def on_startup(app: Application) -> None:
    commands = [
        BotCommand("start", "Show main menu"),
        BotCommand("import_quizzes", "Import quizzes from a file"),
        BotCommand("export_quizzes", "Export quizzes to file"),
        BotCommand("cancel", "Cancel current operation"),
    ]
    await app.bot.set_my_commands(commands)


# -------------------------------
# MAIN: Set Up Handlers and Run the Bot
# -------------------------------
def main():
    # IMPORTANT: Do not hard-code your token in production.
    TOKEN = "7699629853:AAHwJfx-IOBtndlnrTyzJ9G3YKKp-367BhU"
    app = Application.builder().token(TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("export_quizzes", export_quizzes_handler))
    app.add_handler(CommandHandler("import_quizzes", import_quizzes_start))

    import_conversation = ConversationHandler(
        entry_points=[CommandHandler("import_quizzes", import_quizzes_start)],
        states={
            IMPORT_WAITING_FILE: [MessageHandler(filters.Document.ALL, import_quizzes_file)]
        },
        fallbacks=[CommandHandler("cancel", import_quizzes_cancel)],
        allow_reentry=True,
    )
    app.add_handler(import_conversation)

    quiz_creation_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(cr_entry, pattern="^create_quiz$")],
        states={
            CR_QUIZ_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, cr_quiz_name)],
            CR_QUIZ_DESCRIPTION: [
                CommandHandler("skip", cr_skip_description),
                MessageHandler(filters.TEXT & ~filters.COMMAND, cr_quiz_description),
            ],
            CR_PRE_QUESTION: [
                CommandHandler("skip", cr_skip_pre_question),
                MessageHandler(filters.TEXT | filters.PHOTO, cr_pre_question),
            ],
            CR_ADD_QUESTION: [
                MessageHandler(filters.POLL, handle_created_poll),
                CommandHandler("done", cr_finish_questions),
            ],
            CR_QUESTION_IMAGE: [
                CommandHandler("skip", cr_skip_question_image),
                MessageHandler(filters.PHOTO, cr_add_question_image),
            ],
            CR_NEXT_QUESTION: [CallbackQueryHandler(cr_next_question, pattern="^(add_another|finish_questions)$")],
            CR_TIMER: [MessageHandler(filters.TEXT & ~filters.COMMAND, cr_set_timer)],
            CR_SHUFFLE: [CallbackQueryHandler(cr_set_shuffle, pattern="^shuffle_")],
            CR_SUBMIT: [CallbackQueryHandler(cr_set_submit, pattern="^submit_")],
        },
        fallbacks=[CommandHandler("cancel", cr_cancel)],
        allow_reentry=True,
    )
    app.add_handler(quiz_creation_conv)

    quiz_taking_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(tk_entry, pattern="^take_quiz$")],
        states={
            TK_SELECT_QUIZ: [CallbackQueryHandler(tk_select_quiz, pattern="^select_quiz_\\d+$")],
            TK_READY: [CallbackQueryHandler(tk_ready_handler, pattern="^ready$")],
            TK_RUNNING: [CommandHandler("cancel", tk_cancel)],
        },
        fallbacks=[CommandHandler("cancel", tk_cancel)],
        allow_reentry=True,
    )
    app.add_handler(quiz_taking_conv)

    app.add_handler(PollAnswerHandler(handle_poll_answer))

    logging.basicConfig(
        format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO
    )
    logger = logging.getLogger(__name__)
    logger.info("Interactive Quiz Bot is running...")

    loop = asyncio.get_event_loop()
    loop.run_until_complete(app.initialize())
    loop.run_until_complete(on_startup(app))
    app.run_polling()


if __name__ == "__main__":
    main()
